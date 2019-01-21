# -*- coding: utf-8 -*-

from openpyxl import load_workbook

def load_from_file(file, dept_id):
    hardware = []

    try:
        wb2 = load_workbook(file)
    except:
        pass
    else:
        ws = wb2.get_active_sheet()

        for row in xrange(0, ws.get_highest_row()):
            val = ws.cell(row=row, column=35).value
            if val is not None and (val in xrange(10) or val in [11,21,22, 23]):
                if val == 23:
                    val = 3
                elif val == 11:
                    hardware.append(
                        (
                            ws.cell(row=row, column=0).value,
                            ws.cell(row=row, column=2).value,
                            ws.cell(row=row, column=6).value,
                            ws.cell(row=row, column=9).value,
                            ws.cell(row=row, column=12).value,
                            ws.cell(row=row, column=18).value,
                            dept_id,
                            2
                            )
                    )
                    hardware.append(
                        (
                            ws.cell(row=row, column=0).value,
                            ws.cell(row=row, column=2).value,
                            ws.cell(row=row, column=6).value,
                            ws.cell(row=row, column=9).value,
                            ws.cell(row=row, column=12).value,
                            ws.cell(row=row, column=18).value,
                            dept_id,
                            6
                            )
                    )
                else:
                    hardware.append(
                        (
                            ws.cell(row=row, column=0).value,
                            ws.cell(row=row, column=2).value,
                            ws.cell(row=row, column=6).value,
                            ws.cell(row=row, column=9).value,
                            ws.cell(row=row, column=12).value,
                            ws.cell(row=row, column=18).value,
                            dept_id,
                            val #ws.cell(row=row, column=35).value
                            )
                )
    return hardware


if __name__ == '__main__':
    print load_from_file('5.xlsx', 66)
